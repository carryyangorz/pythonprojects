import openai
from openpyxl import Workbook
from openpyxl import load_workbook
from time import sleep
import os
openai.api_key = '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'
print('please input interval time:')
sleeptime=int(input())
#print('please input engine name:')
#enginee=input()
enginels=['text-davinci-001',

'text-curie-001'
          ]
wb=load_workbook('chat.xlsx')
question=[]
ws=wb.active
for i in range(2,ws.max_row+1):
    if ws.cell(row=i,column=2).value!=None:
        print(ws.cell(row=i,column=1).value + "   pass")
        continue
    if ws.cell(row=i,column=1).value==None:
        break

    print("==============================")
    print(ws.cell(row=i,column=1).value)
    question=(ws.cell(row=i,column=1).value)
    try:
        responses=openai.Completion.create(
        engine=enginels[0],
        prompt=question,
        max_tokens=2000,
        stop=None,
        n=1
    )
    except:
        responses=openai.Completion.create(
        engine=enginels[1],
        prompt=question,
        max_tokens=2000,
        stop=None,
        n=1
    )
    sleep(sleeptime)
    print(responses.choices[0].text)
    print(type(responses.choices[0].text))
    print("==============================")

    ws.cell(row=i,column=2).value=(responses.choices[0].text).replace('\n','')
    wb.save('chat.xlsx')
wb.close()
os.system('pause')
