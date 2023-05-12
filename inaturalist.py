import requests
import os
import time
import sys
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import ChromeOptions
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from threading import Timer
import multiprocessing
capa = DesiredCapabilities.CHROME
capa["pageLoadStrategy"] = "eager"
option=ChromeOptions()
option.add_argument('--headless')
option.add_argument('--no-sandbox')
option.add_argument('log-level=3') #INFO = 0 WARNING = 1 LOG_ERROR = 2 LOG_FATAL = 3 default is 0
option.add_experimental_option('excludeSwitches',['enable-automation'])
# browser=webdriver.Chrome(options=option,desired_capabilities=capa)
worldls=['https://www.inaturalist.org/taxa/1-Animalia',
'https://www.inaturalist.org/taxa/47126-Plantae',
'https://www.inaturalist.org/taxa/47170-Fungi',
'https://www.inaturalist.org/taxa/47686-Protozoa',
'https://www.inaturalist.org/taxa/131236-Viruses',
'https://www.inaturalist.org/taxa/48222-Chromista',
'https://www.inaturalist.org/taxa/67333-Bacteria']
headers={
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36'
}

def get(baseurl):
    browser.get(baseurl)
    sleep(1)
    tab=browser.find_element_by_css_selector('div.TaxonPageTabs')
    a=tab.find_element_by_css_selector('div.col-xs-12')
    a.find_elements_by_css_selector('li')[3].click()
    sleep(1)
    b=browser.find_element_by_css_selector('div.TaxonomicBranch')
    c=b.find_elements_by_css_selector('li.all-shown.tabular')
    i=0
    ls=[]
    for item in c:
        # print(item.text)
        count=item.text.split(' ')[-1]

        if count != '0':
            
            print(str(i)+'---'+item.text)
            ls.append(item)
            i=i+1
    print('请选择： 输入数字并回车')
    num=int(input())
    url=ls[num].find_elements_by_css_selector('a')[0].get_attribute('href')
    print(url)
    getsub(url)

def getsub(url):
    browser.get(url)
    sleep(1)
    tab=browser.find_element_by_css_selector('div.TaxonPageTabs')
    d=tab.find_element_by_css_selector('div.col-xs-12')
    d.find_elements_by_css_selector('li')[3].click()
    sleep(1)
    e=browser.find_element_by_css_selector('div.TaxonomicBranch')
    f=e.find_elements_by_css_selector('li.all-shown.tabular')
    i=0
    ls=[]
    for item in f:
        # print(item.text)
        count=item.text.split(' ')[-1]

        if count != '0':
            
            print(str(i)+'---'+item.text)
            ls.append(item)
            i=i+1
    print('请选择： 输入数字并回车')
    num=int(input())
    url=ls[num].find_elements_by_css_selector('a')[0].get_attribute('href')
    print(url)
    getfinal(url)

def getfinal(url):
    browser.get(url)
    sleep(1)
    tab=browser.find_element_by_css_selector('div.TaxonPageTabs')
    d=tab.find_element_by_css_selector('div.col-xs-12')
    d.find_elements_by_css_selector('li')[3].click()
    sleep(1)
    # e=browser.find_element_by_css_selector('div.TaxonomicBranch')
    e=browser.find_element_by_css_selector('ul.plain.taxonomy')
    f=e.find_elements_by_css_selector('li.all-shown.tabular')
    i=0
    uls=[]
    nls=[]
    for item in f:
        # print(item.text)
        count=item.text.split(' ')[-1]
        if count != '0':
            print(str(i)+'---'+item.text)
            url=item.find_elements_by_css_selector('a')[-1].get_attribute('href')
            name=re.findall('[a-zA-Z]{2,100}',item.text)[0]
            # print('ssssssssssssssss'+ '   '+name)
            # aa=[url,name]
            uls.append(url)
            nls.append(name)
            i=i+1
    for i in range(0,len(uls)):
        download(uls[i],nls[i])
        # download(item)

def download(aurl,name):
    # url=item[0]
    # name=url[1]
    # print(url)
    # print(name)
    id=re.findall('taxon_id=(.*)&place',aurl)[0]
    url='https://api.inaturalist.org/v1/observations?verifiable=any&order_by=observations.id&order=desc&page=1&spam=false&taxon_id='+id+'&locale=zh-CN&per_page=200'
    browser.get(url)
    ls=[]
    aaa=browser.find_element_by_css_selector('pre').text
    total=re.findall('total_results":(.*),"page',aaa)[0]
    # print(total)
    pages=int(total)//200+2
    for i in range(1,pages):
        print(i)
        url='https://api.inaturalist.org/v1/observations?verifiable=any&order_by=observations.id&order=desc&page='+str(i)+'&spam=false&taxon_id='+id+'&locale=zh-CN&per_page=200'
        browser.get(url)
        aaa=browser.find_element_by_css_selector('pre').text
        b=re.findall('square_url":"(.*?)","attribution',aaa)
        print(len(b))
        for item in b:
            if "inaturalist" in item:
                ls.append(item)
        i=i+1
    lss=list(set(ls))
    print(len(lss))

    print(name+'   total:   '+str(len(lss)))
    i=1
    if not os.path.exists(name+'\\'):
        os.mkdir(name+'\\')
    for item in lss:
        # pid=re.findall('photos/(.*)/large')
        picurl=item.replace('square','large')
        r=requests.get(picurl,headers=headers)
        f=open(name+'\\'+name+'_'+str(i)+'.jpg','wb')
        f.write(r.content)
        f.close()
        f=open('url.txt','a')
        f.write('\n'+name+'_'+str(i)+'\t'+aurl)
        f.close()
        print(name+'   '+str(i))
        i=i+1
def savedata(data):
    countryname=data[0][0]
    species=data[0][1]
    data.remove(data[0])
    wb=load_workbook(countryname+'_'+species+'.xlsx')
    ws=wb.active
    for item in data:
        ws.append(item)
    wb.save(countryname+'_'+species+'.xlsx')
def getlist(species,countryid,countryname):
    
    if not os.path.exists(countryname+'_'+species+'.xlsx'):
        # f=open(countryname+'_'+species+'.xls','a',encoding='utf-8')
        # f.write('Latin'+'\t')
        # f.write('Chinese Name'+'\t')
        # f.write('photo'+'\n')
        # f.close()
        wb=Workbook()
        ws=wb.active
        ws.append(['Latin','Chinese Name','photo'])
        wb.save(countryname+'_'+species+'.xlsx')
    else:
        print(countryname+'_'+species+'.xlsx'+' already exists')
        return

    for page in range(1,2000000):
        data=[[countryname,species]]
        speciesurl='https://api.inaturalist.org/v1/observations/species_counts?verifiable=true&spam=false&place_id='+countryid+'&iconic_taxa%5B%5D='+species+'&locale=zh-CN&page='+str(page)+'&per_page=200'
        r=requests.get(speciesurl,headers=headers)
        #print('saving ' +countryname+'  '+species+' page :'+str(page))
        aaa=re.findall('("count":.*?)}}',r.text)
        if(len(aaa)==0):
            print('over')
            return
        print('saving ' +countryname+'  '+species+' page :'+str(page)+' number: '+str(len(aaa)))
        for item in aaa:
            latin=re.findall('name":"(.*?)","rank',item)

            chinese=re.findall('preferred_common_name":"(.*?)"',item)
            photo=re.findall('"count":(.*?),"taxon',item)

            if len(latin)==0:
                latin="NULL"
            else:
                latin=latin[0]
            
            if len(chinese)==0:
                chinese='NULL'
            else:
                chinese=chinese[0]

            if len(photo)==0:
                photo='NULL'
            else:
                photo=photo[0]
            data.append([latin,chinese,photo])
        t=multiprocessing.Process(target=savedata,args=(data,))
        t.start()
        #t.join()
            # f=open(countryname+'_'+species+'.xls','a',encoding='utf-8')
            # try:
            #     f.write(latin[0]+'\t')
            #     print(latin[0])
            # except:
            #     f.write('NULL'+'\t')
            # try:
            #     f.write(chinese[0]+'\t')
            #     print(chinese[0])
            # except:
            #     f.write('NULL'+'\t')
            # try:
            #     f.write(photo[0]+'\n')
            #     print(photo[0])
            # except:
            #     f.write('NULL'+'\n')
            # f.close()
        

speciesls=['Aves',    'Amphibia',   'Reptilia',   'Mammalia',   'Actinopterygii',   'Mollusca',
    'Arachnida',   'Insecta', 'Plantae',  'Fungi',  'Protozoa',  'unknown']
if __name__ == "__main__":
    
    multiprocessing.freeze_support()
    print('请输入国家名称：')
    countryinput=input()
    countryurl='https://api.inaturalist.org/v1/search?callback=placeAutocompleteCallback&q='+countryinput+'&per_page=10&sources=places'
    r=requests.get(countryurl,headers=headers)
    aaa=re.findall('id":(.*?),"uuid',r.text)
    
    countryname=re.findall('"matches":\["(.*?)"',r.text)[0]
    if len(aaa)==0:
        print('没有找到，请重新输入')
        sys.exit()
    else:
        print(countryname)
        counrtyid=aaa[0]
    print('请选择界：0.Aves    1.Amphibia   2.Reptilia   3.Mammalia   4.Actinopterygii   5.Mollusca')
    print('         6.Arachnida   7.Insecta  8.Plantae  9.Fungi  10.Protozoa  11.unknown')
    species=speciesls[int(input())]
    getlist(species,counrtyid,countryname)
    # # baseurl=worldls[int(world)]
    # baseurl='https://www.inaturalist.org/taxa/67333-Bacteria'
    # get(baseurl)
