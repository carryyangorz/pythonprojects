import requests
import os
import time
import re
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import ChromeOptions
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from threading import Timer
import threading
capa = DesiredCapabilities.CHROME
capa["pageLoadStrategy"] = "eager"
option=ChromeOptions()
# option.add_argument('--headless')
option.add_argument('--no-sandbox')
option.add_argument('log-level=3') #INFO = 0 WARNING = 1 LOG_ERROR = 2 LOG_FATAL = 3 default is 0
option.add_experimental_option('excludeSwitches',['enable-automation'])
# option.add_argument('--blink-settings=imagesEnabled=false')

# option.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
import multiprocessing
from multiprocessing import Process,Lock
# import multiprocessing_win
headers={
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36'
        }
baseurl='https://api.plantnet.org/v1/projects/'
countrycodels=['the-plant-list','useful','weeds','invasion','prota','prosea',
       'weurope','canada','namerica','central-america','antilles','colombia','guyane','brazil','lapaz','martinique',
               'afn','aft','reunion','maurice','comores','medor','malaysia','japan',
           'nepal','endemia','hawai','polynesiefr']
countryls=['WorldFlora','UsefulPlants','Weeds','InvasivePlants','UsefulPlantsTropicalAfrica','UsefulPlantsAsia',
       'WesternEurope','Canada','USA','CentralAmerica','Caribbean','Colombia','Amazonia','Brazil','TropicalAndes','Martinique',
           'NorthAfrica','TropicalAfrica','Reunion','Mauritius','ComoroIslands','EasternMediterranean','Malaysia','Japan',
           'Nepal','NewCaledonia','Hawaii','FrenchPolynesia']
def get(i):
       if not os.path.exists(countryls[i]+'.xls'):
              wb=Workbook()
              ws=wb.active
              ws.append(['Species','common name','photo','observation','Family'])
              wb.save(countryls[i]+'.xlsx')

       else:
              print('please remove '+countryls[i]+'.xls'+' and then try again')
              return
       for page in range(10000):
              time1=time.mktime(time.localtime())
              data=[[countryls[i]]]
              url2=baseurl+countrycodels[i] + '/species?pageSize=400&page='+str(page)+'&lang=en&sortBy=images_count&sortOrder=desc&illustratedOnly=true'
              browser.get(url2)
              r=browser.find_element(By.CSS_SELECTOR,'pre').text
              # r=requests.get(url2,headers=headers)
              b=re.findall('name(.*?)searchTerms"',r)
              time2=time.mktime(time.localtime())
              print('---------------------------')
              print(countryls[i]+'  page NO.' + str(page)+' lenth:'+str(len(b)))
              if (len(b)==0):
                     print(countryls[i]+'    over')
                     return
              for item in b:
                     name=re.findall('":"(.*?)","auth',item)
                     auth=re.findall('author":"(.*?)"',item)
                     commonname=re.findall('commonNames":\["(.*?)"',item)
                     photo=re.findall('imagesCount":(.*?),"',item)
                     observation=re.findall('observationsCount":(.*?),"',item)
                     family=re.findall('family":"(.*?)"',item)
                     
                     if len(name)==0:
                            name='NULL'
                     else:
                            name=name[0]
                     
                     if len(auth) == 0:
                            auth='NULL'
                     else:
                            auth=auth[0]

                     if len(commonname)==0:
                            commonname='NULL'
                     else:
                            commonname=commonname[0]
                     
                     if len(photo)==0:
                            photo='NULL'
                     else:
                            photo=photo[0]
                     
                     if len(observation)==0:
                            observation='NULL'
                     else:
                            observation=observation[0]
                     
                     if len(family)==0:
                            family='NULL'
                     else:
                            family=family[0]
                     name=name+' '+auth
                     data.append([name,commonname,photo,observation,family])
              t=multiprocessing.Process(target=savedata,args=(data,))
              t.start()
                     #    wb=load_workbook(countryls[i]+'.xlsx')
                     #    ws=wb.active
                     #    ws.append([name+' '+auth,commonname,photo,observation,family])
                     #    wb.save(countryls[i]+'.xlsx')
              print('timecost ' +str(time2-time1))
              print('---------------------------')
specie_type_ls=['Flower',
'Leaf',
'Fruit',
'Bark',
'Habit',
'Other'
]
def savedata(data):
    bookname=data[0][0]
    data.remove(data[0])
    wb=load_workbook(bookname+'.xlsx')
    ws=wb.active
    for item in data:
        ws.append(item)
    wb.save(bookname+'.xlsx')

def savedata_2(data):
    name=data[0][0]
    specie_type=data[0][1]
    data.remove(data[0])
    wb=load_workbook(name+"//"+name+"_"+specie_type+'.xlsx')
    ws=wb.active
    for item in data:
        ws.append(item)
    wb.save(name+"//"+name+"_"+specie_type+'.xlsx')
def download_species(name,num,typeindex):
    if not os.path.exists(name+'//'):
       os.makedirs(name+'//')
    else:
          print(name+' already exists,pass')
          return
    print(name+' searching...')
    itemnum=0
    nameurl='https://api.plantnet.org/v1/projects/namerica/species?pageSize=50&page=0&lang=en&search='+name+'&sortBy=images_count&sortOrder=desc'
    r=requests.get(nameurl,headers=headers)
    auth=re.findall('"author":"(.*)","fam',r.text)[0]
    searchname=name.replace(' ','%20')
    if auth!='':
       searchname=searchname+'%20'+auth
    url='https://identify.plantnet.org/namerica/species/'+searchname+'/data'
    print(searchname)
    browser.get(url)
    js="var q=document.documentElement.scrollTop=100000"
    browser.execute_script(js)
    sleep(100)
    try:
       father=browser.find_element(By.CSS_SELECTOR,'nav.card-header')
    except:
          print(name+' not found please try again later')
          return
    if not os.path.exists(name+'//'+name+'_'+specie_type_ls[typeindex]+'.xlsx'):
        wb=Workbook()
        ws=wb.active
        ws.append(['图片名','来源','网址'])
        wb.save(name+'//'+name+'_'+specie_type_ls[typeindex]+'.xlsx')
    for i in range(6):
       # i=species_t%6
       type=father.find_elements(By.CSS_SELECTOR,'li.nav-item')
       print(name+'  types: '+str(len(type)))
       if i>=len(type):
             break
       b=type[i].find_element(By.CSS_SELECTOR,'a')
       species_type=b.find_element(By.CSS_SELECTOR,'img').get_attribute('alt').capitalize()
       if typeindex!=9:
             if species_type!=specie_type_ls[typeindex]:
              #      print('ssssssssssssssssssssssssssssssssssssssssssssssssssssss')
                   continue   
       print('--------------------------------------------------')
       b.send_keys(Keys.ENTER)
       sleep(4)
       # species_type=specie_type_ls[i]
       while True:
              a=browser.find_elements(By.CSS_SELECTOR,"img.img-fluid")
              js="var q=document.documentElement.scrollTop=100000"
              browser.execute_script(js)
              sleep(10)
              # hei=0
              # for y in range(5):
              #       hei+=1000
              #       browser.execute_script(f'window.scrollto(0,{hei})')
              # for item in a:
              #        item.send_keys(Keys.ENTER)
              #        sleep(10)
              #        info=browser.find_element(By.CSS_SELECTOR,'header.modal-header')
              #        print(info)
              #        sleep(5)
              if itemnum==len(a):
                     print('not increasing')
                     break
              itemnum=len(a)
              # print(itemnum)
              if itemnum>num:
                     break

       print(name+' '+species_type+' total:'+str(itemnum))
       ls1=[]
       ls2=[]
       ls3=[]
       j=0
       for i in range(itemnum):
              if i>5000:
                    break
              picurl=a[i].get_attribute('src')
              if j>num-1:
                    break
              if picurl==None:
                     # print('-1')
                     picurl=a[i].get_attribute('data-src')
                     if picurl==None:
                            print('-2')
                            continue
                     else:
                           pass
                     #       print('ok  '+picurl)
              j=j+1
              picurl=picurl.replace('/s/','/o/')
              ls1.append(picurl)
              picid=re.findall('/o/(.*)',picurl)[0]
              ls2.append(picid)
              ls3.append(name+'_'+species_type+'_'+picid+'.jpg')
       # if species_t>5:
       data=[[name,species_type]]
       for i in range (len(ls3)):
              data.append([ls3[i],"",ls1[i]])
       t=multiprocessing.Process(target=savedata_2,args=(data,))
       t.start()
              # #  print(picid)
              # r=requests.get(picurl,headers=headers)
              # r.raise_for_status()
              # #  if os.path.exists(name+'\\'+name+'_'+picid+'.jpg'):
              # #        continue
              # f=open(name+'//'+name+'_'+species_type+'_'+picid+'.jpg','wb')
              # f.write(r.content)
              # print(name+'  '+species_type+'   +1')
              # f.close()
       ppls=[]
       i=0
       flag=1
       while flag:
              for j in range(40):
                     if i>len(ls1)-1:
                            flag=0
                            break
                     data=[ls1[i],ls2[i],name,species_type]
                     pp=Process(target=download,args=(data,))
                     i=i+1
                     # if i>len(bb):
                     #         break
                     pp.start()
              # ppls.append(pp)
       #      f=open(target+'\\'+target+'.txt','a')
       #      f.write('\n'+als[i]+'_'+str(i)+'.jpg''\t'+bls[i]+'\t'+bls[i].replace('midthumb','smthumb')+'\t'+cls[i]+'\t'+dls[i])
       #      f.close()
       #  for thread in ppls:
       #      thread.join()
def download(data):
    
    picurl=data[0]
    picid=data[1]
    name=data[2]
    species_type=data[3]
#     print(picurl)
    if os.path.exists(name+'\\'+name+'_'+species_type+'_'+picid+'.jpg'):
        print(name+'_'+species_type+'_'+picid+'.jpg already exists')
        return
    
    try:
        r=requests.get(picurl,headers=headers)#,verify=False)
    except:
        print('error')
        # continue
        return
    f=open(name+'\\'+name+'_'+species_type+'_'+picid+'.jpg','wb')
    # r.raise_for_status()
    f.write(r.content)
    # i=i+1
    print(name+' '+species_type+'  ''+1')
    f.close()
      
if __name__ == "__main__":
    multiprocessing.freeze_support()
    browser=webdriver.Chrome(options=option,desired_capabilities=capa)
    browser.implicitly_wait(6)
    #1.  get list workbook
#     print('请输入要下载的序号，如0，1，2,输入999下载全部')
#     i=0
#     for item in countryls:
#            print(str(i)+'.'+item+'\t')
#            i=i+1
#     choice=input()
#     if(choice=='999'):
           
#         for i in range(len(countrycodels)):
#                 get(i)
#     else:
#            get(int(choice))
       #2. download
    print('请输入下载的品种编号：0.Flower 1.Leaf 2.Fruit 3.Bark 4.Habit 5.Other 9.All')
    typeindex=int(input())
    print('请输入每个品种下载个数：')
    num=int(input())
#     num=40
    f = open("plantnet.txt",encoding='utf-8')
    line = f.readline()
    ls=[]
    # errorls=[]
    while line:
        ls.append(line.replace('\n',''))
        line = f.readline()
    f.close()
    
    for item in ls:
           download_species(item,num,typeindex)
#     download_species(name,num)
    os.system ("pause")