import requests
import os
import time
import re
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import ChromeOptions
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from threading import Timer
import threading
capa = DesiredCapabilities.CHROME
capa["pageLoadStrategy"] = "eager"
option=ChromeOptions()
# option.add_argument('--headless')
option.add_argument('--no-sandbox')
option.add_argument('log-level=3') #INFO = 0 WARNING = 1 LOG_ERROR = 2 LOG_FATAL = 3 default is 0
option.add_experimental_option('excludeSwitches',['enable-automation'])
# option.add_argument('--blink-settings=imagesEnabled=false')

# option.add_experimental_option("debuggerAddress", "127.0.0.1:9527")
import multiprocessing
from multiprocessing import Process,Lock
# import multiprocessing_win
headers={
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36'
        }
baseurl='https://api.plantnet.org/v1/projects/'
countrycodels=['the-plant-list','useful','weeds','invasion','prota','prosea',
       'weurope','canada','namerica','central-america','antilles','colombia','guyane','brazil','lapaz','martinique',
               'afn','aft','reunion','maurice','comores','medor','malaysia','japan',
           'nepal','endemia','hawai','polynesiefr']
countryls=['WorldFlora','UsefulPlants','Weeds','InvasivePlants','UsefulPlantsTropicalAfrica','UsefulPlantsAsia',
       'WesternEurope','Canada','USA','CentralAmerica','Caribbean','Colombia','Amazonia','Brazil','TropicalAndes','Martinique',
           'NorthAfrica','TropicalAfrica','Reunion','Mauritius','ComoroIslands','EasternMediterranean','Malaysia','Japan',
           'Nepal','NewCaledonia','Hawaii','FrenchPolynesia']

def savedata(data):
    bookname=data[0][0]
    data.remove(data[0])
    wb=load_workbook(bookname+'.xlsx')
    ws=wb.active
    for item in data:
        ws.append(item)
    wb.save(bookname+'.xlsx')

def download_species(name):
    if os.path.exists(name+'.xlsx'):
        print(name+' already exists, pass')
        return

    nameurl='https://api.plantnet.org/v1/projects/the-plant-list/species?pageSize=50&page=0&lang=en&search='+name+'&sortBy=images_count&sortOrder=desc'
    r=requests.get(nameurl,headers=headers)
    try:
        auth=re.findall('"author":"(.*?)","fam',r.text)[0]
    except:
        print(name+' not found')
        return

    searchname=name.replace(' ','%20')
    if auth!='':
       searchname=searchname+'%20'+auth
    url='https://identify.plantnet.org/the-plant-list/species/'+searchname+'/data'
    print(url)
    try:
       r=requests.get(url,headers=headers)

       a=r.text.replace('\n','')
       b=re.findall('text-secondary mt-n5">(.*?)</h5>',a)[0]
       b=b.replace('\t','')
       # print(b)
       # d=re.findall('commonNames:\[(.*?)\],synonyms',a)[0]
       # d=d.replace('\t','')
       # d=d.replace('"','')
       # print(d)
       # c=d.split(',')

       url2='https://api.plantnet.org/v1/projects/the-plant-list/species/'+searchname+'/vernaculars?lang=en'
       r=requests.get(url2,headers=headers)
       c=re.findall('English","terms":\[(.*?)\]}',r.text)[0]
       c=c.split(',')
    except:
        print(name+' not found2')
        return
#     print(c)
    if len(c)!=0:
       wb=Workbook()
       ws=wb.active
       ws.append(['搜索名','网址','学名','别名','英文名'])
    for item in c:
        item=item.replace('"','')
        ws.append([name,url,searchname.replace('%20',' '),b,item])
        print(name+' +1')

    wb.save(name+'.xlsx')
    wb.close()
        

      
if __name__ == "__main__":
    multiprocessing.freeze_support()

    f = open("plantnet.txt",encoding='utf-8')
    line = f.readline()
    ls=[]
    # errorls=[]
    while line:
        ls.append(line.replace('\n',''))
        line = f.readline()
    f.close()
    
    for item in ls:
           download_species(item)
#     name='Paulownia tomentosa'
#     download_species(name)
    os.system ("pause")